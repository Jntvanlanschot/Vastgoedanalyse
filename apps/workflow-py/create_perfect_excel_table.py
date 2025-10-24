#!/usr/bin/env python3
"""
Create Excel table from PERFECT top 15 data.
"""

import pandas as pd
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

def create_perfect_excel_table():
    """Create a beautiful Excel table from perfect top 15 data."""
    
    # Load the perfect comparison data
    df = pd.read_csv('outputs/top15_perfect_matches.csv')
    
    logger.info(f"Loaded {len(df)} houses from perfect top 15 data")
    
    excel_data = []
    for i, row in df.iterrows():
        # Create address from Funda data
        street = row.get('address/street_name', '')
        house_num = row.get('address/house_number', '')
        house_suffix = row.get('address/house_number_suffix', '')
        postal_code = row.get('address/postal_code', '')
        city = row.get('address/city', '')
        
        address = f"{street} {house_num}{house_suffix}, {postal_code}, {city}"
        
        # Calculate price per m² from Realworks data
        sale_price = row.get('rw_sale_price', 0)
        area_m2 = row.get('rw_area_m2', 0)
        price_per_m2 = sale_price / area_m2 if area_m2 > 0 else 0
        
        excel_data.append({
            '#': i + 1,
            'Adres': address,
            'Verkoopprijs': f"€{sale_price:,.0f}" if pd.notna(sale_price) and sale_price > 0 else 'Onbekend',
            'Woonoppervlakte (m²)': area_m2 if pd.notna(area_m2) and area_m2 > 0 else 'Onbekend',
            'Kamers': int(row['rw_rooms']) if pd.notna(row['rw_rooms']) else 'Onbekend',
            'Slaapkamers': int(row['rw_bedrooms']) if pd.notna(row['rw_bedrooms']) else 'Onbekend',
            'Badkamers': int(row['rw_bathrooms']) if pd.notna(row['rw_bathrooms']) else 'Onbekend',
            'Bouwjaar': int(row['rw_year_built']) if pd.notna(row['rw_year_built']) else 'Onbekend',
            'Type': row['rw_type'] if pd.notna(row['rw_type']) else 'Onbekend',
            'Subtype': row['rw_subtype'] if pd.notna(row['rw_subtype']) else 'Onbekend',
            'Tuin': 'Ja' if row.get('rw_has_garden') else 'Nee',
            'Balkon': 'Ja' if row.get('rw_has_balcony') else 'Nee',
            'Terras': 'Ja' if row.get('rw_has_terrace') else 'Nee',
            'Zon oriëntatie': row['rw_outdoor_text'] if pd.notna(row['rw_outdoor_text']) else 'Onbekend',
            'Onderhoud binnen': row['rw_maintenance_inside'] if pd.notna(row['rw_maintenance_inside']) else 'Onbekend',
            'Onderhoud buiten': row['rw_maintenance_outside'] if pd.notna(row['rw_maintenance_outside']) else 'Onbekend',
            'Energielabel': row['rw_energy_label'] if pd.notna(row['rw_energy_label']) else 'ONBEKEND',
            'Prijs per m²': f"€{price_per_m2:,.0f}" if price_per_m2 > 0 else 'Onbekend',
            'Score': f"{row['final_score']:.2f}",
            'Match Type': row.get('match_type', 'unknown'),
            'Funda Link': f"https://www.funda.nl{row['object_detail_page_relative_url']}" if pd.notna(row['object_detail_page_relative_url']) else 'Geen link'
        })
    
    # Create DataFrame
    excel_df = pd.DataFrame(excel_data)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Top 15 Perfecte Woningen"
    
    # Add data to worksheet
    for r in dataframe_to_rows(excel_df, index=False, header=True):
        ws.append(r)
    
    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set column widths
    column_widths = {
        'A': 5,   # #
        'B': 35,  # Adres
        'C': 15,  # Verkoopprijs
        'D': 18,  # Woonoppervlakte
        'E': 8,   # Kamers
        'F': 12,  # Slaapkamers
        'G': 10,  # Badkamers
        'H': 10,  # Bouwjaar
        'I': 12,  # Type
        'J': 15,  # Subtype
        'K': 6,   # Tuin
        'L': 8,   # Balkon
        'M': 8,   # Terras
        'N': 15,  # Zon oriëntatie
        'O': 15,  # Onderhoud binnen
        'P': 15,  # Onderhoud buiten
        'Q': 12,  # Energielabel
        'R': 12,  # Prijs per m²
        'S': 8,   # Score
        'T': 10,  # Match Type
        'U': 50   # Funda Link
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add borders to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Save Excel file
    output_file = 'outputs/top15_perfecte_woningen_tabel.xlsx'
    wb.save(output_file)
    logger.info(f"Perfect Excel table created: {output_file}")
    
    # Print summary
    print("\n=== TOP 15 PERFECTE WONINGEN TABEL ===")
    print(f"Bestand: {output_file}")
    print(f"Aantal woningen: {len(excel_df)}")
    
    # Calculate average price per m²
    price_values = excel_df['Prijs per m²'].str.replace('€', '').str.replace(',', '').str.replace('Onbekend', '0')
    price_values = pd.to_numeric(price_values, errors='coerce')
    avg_price = price_values[price_values > 0].mean()
    print(f"Gemiddelde prijs per m²: €{avg_price:,.0f}")
    
    return output_file

if __name__ == "__main__":
    create_perfect_excel_table()

